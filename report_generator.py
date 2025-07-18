import streamlit as st
import pandas as pd
import os
from datetime import datetime, date, timedelta
import io
import zipfile
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import warnings
import numpy as np
import glob

warnings.filterwarnings("ignore", category=UserWarning, module='openpyxl')

from data_ingestion import load_data
from data_cleaning import clean_and_filter_data, ASSERTIVA_ESSENTIAL_COLS, LEMIT_ESSENTIAL_COLS
from create_pdf import create_pdf_robust

# --- Configurações e Lógica para o Divisor de Listas ---
from config import CONSULTORES, EQUIPES, CONSULTOR_PARA_EQUIPE

# Cores para o Excel (RGB para OpenPyXL)
COLOR_LIGHT_BLUE = "E0EBFB"
COLOR_WHITE = "FFFFFF"

def clean_phone_number(number_str):
    """Limpa e valida um número de telefone, retornando NaN se inválido."""
    if pd.isna(number_str) or number_str == '':
        return np.nan
    cleaned = ''.join(filter(str.isdigit, str(number_str)))
    if len(cleaned) < 8:
        return np.nan
    return cleaned

def proximo_dia_util(data_atual):
    proximo_dia = data_atual + timedelta(days=1)
    while proximo_dia.weekday() >= 5:
        proximo_dia += timedelta(days=1)
    return proximo_dia

def formatar_planilha(writer, df):
    workbook = writer.book
    worksheet = writer.sheets['Leads']
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    cols_to_center = ["1º Contato", "2º Contato", "3º Contato", "Atend. Lig.(S/N)", "Visita Marc.(S/N)"]

    for col_idx, column_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=2, column=col_idx)
        cell.border = thin_border
        cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
        cell.font = Font(bold=True)
        if column_name in cols_to_center:
            cell.alignment = center_align

    for row_idx in range(3, len(df) + 3):
        fill_color = COLOR_LIGHT_BLUE if (row_idx - 3) % 2 == 0 else COLOR_WHITE
        for col_idx, column_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            if column_name in cols_to_center:
                cell.alignment = center_align

    for col_idx, column_name in enumerate(df.columns, 1):
        max_length = 0
        header_cell = worksheet.cell(row=2, column=col_idx)
        if header_cell.value:
            max_length = len(str(header_cell.value))
        for row_idx in range(3, len(df) + 3):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[header_cell.column_letter].width = adjusted_width

def gerar_excel_em_memoria(df_lote, consultor, data):
    output = io.BytesIO()
    primeiro_nome = consultor.split(' ')[0]
    data_formatada_cabecalho = data.strftime('%d/%m')
    cabecalho_texto = f"Leads Automoveis - {primeiro_nome} {data_formatada_cabecalho}"

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_lote.to_excel(writer, index=False, startrow=1, sheet_name='Leads')
        workbook = writer.book
        worksheet = writer.sheets['Leads']
        num_colunas = len(df_lote.columns)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_colunas)
        cell = worksheet['A1']
        cell.value = cabecalho_texto
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        formatar_planilha(writer, df_lote)
    
    output.seek(0)
    return output

def aba_higienizacao():
    st.header("Higienização e Geração de Relatórios")
    uploaded_file = st.file_uploader("Faça upload do arquivo CSV", type=["csv"], key="higienizacao_uploader")

    if uploaded_file:
        # Reinicia o estado se um novo arquivo for carregado
        if st.session_state.get('last_uploaded_file') != uploaded_file.name:
            for key in list(st.session_state.keys()):
                if key.startswith('map_higienizacao_') or key in ['df_clean', 'df_export', 'processing_complete', 'structure_type']:
                    del st.session_state[key]
            st.session_state.last_uploaded_file = uploaded_file.name

        temp_path = "temp_uploaded.csv"
        with open(temp_path, "wb") as f:
            f.write(uploaded_file.read())
        df_raw, err = load_data(temp_path)
        if err:
            st.error(err)
            return

        st.subheader("Pré-visualização dos Dados Brutos")
        st.dataframe(df_raw.head(50))

        # --- Detecção Automática da Estrutura ---
        df_cols_set = set(df_raw.columns)
        # Critérios para Lemit: presença de colunas específicas ou maior correspondência com colunas essenciais Lemit
        is_lemit = any(col in df_cols_set for col in ['RAZAO_SOCIAL', 'FULL-LOGRADOURO', 'POSSUI-WHATSAPP']) or \
                   len(df_cols_set.intersection(LEMIT_ESSENTIAL_COLS)) > len(df_cols_set.intersection(ASSERTIVA_ESSENTIAL_COLS))
        
        structure_type = "Lemit" if is_lemit else "Assertiva"
        st.session_state.structure_type = structure_type

        st.subheader("Mapeamento de Colunas")
        st.info(f"**Estrutura detectada:** {structure_type}. Verifique o mapeamento automático abaixo.")

        essential_cols = LEMIT_ESSENTIAL_COLS if structure_type == "Lemit" else ASSERTIVA_ESSENTIAL_COLS
        
        # --- Mapeamento Automático de Colunas ---
        df_raw_cols = df_raw.columns.tolist()
        # Mapeamento normalizado para busca case-insensitive e ignorando separadores
        df_cols_lower_map = {c.lower().replace('_', '').replace('-', ''): c for c in reversed(df_raw_cols)}

        user_col_mapping = {}
        for col in essential_cols:
            default_selection = ''
            # Normaliza a coluna essencial para busca
            search_key = col.lower().replace('_', '').replace('-', '')
            
            if search_key in df_cols_lower_map:
                default_selection = df_cols_lower_map[search_key]
            
            try:
                default_index = df_raw_cols.index(default_selection) + 1 if default_selection else 0
            except ValueError:
                default_index = 0

            selected_col = st.selectbox(
                f"Coluna para '{col}'",
                options=[''] + df_raw_cols,
                index=default_index,
                key=f"map_higienizacao_{col}"
            )
            user_col_mapping[col] = selected_col

        if st.button("Processar Dados"):
            with st.spinner("Processando... Aguarde."):
                df_mapped = df_raw.copy()
                rename_dict = {v: k for k, v in user_col_mapping.items() if v}
                df_mapped.rename(columns=rename_dict, inplace=True)

                df_clean, missing, structure = clean_and_filter_data(df_mapped)
                
                st.session_state.df_clean = df_clean
                st.session_state.missing_cols = missing
                st.session_state.structure_type = structure
                st.session_state.processing_complete = True

        if st.session_state.get('processing_complete', False):
            st.subheader("Dados Higienizados")
            st.success(f"Planilha {st.session_state.structure_type} Detectada")
            st.dataframe(st.session_state.df_clean.head(50))
            st.info(f"Linhas finais: {len(st.session_state.df_clean)}")
            if st.session_state.missing_cols:
                st.warning(f"Colunas essenciais ausentes: {', '.join(st.session_state.missing_cols)}")

            df_export = st.session_state.df_clean.drop(columns=['Distancia'], errors='ignore')

            # Limpeza e formatação de colunas para exportação
            for col_celular in ['SOCIO1Celular1', 'SOCIO1Celular2']:
                if col_celular in df_export.columns:
                    df_export[col_celular] = pd.to_numeric(df_export[col_celular], errors='coerce')
                    df_export[col_celular] = df_export[col_celular].astype('Int64').astype(str)
                    df_export[col_celular] = df_export[col_celular].apply(lambda x: 
                        f'+55{str(x).replace("+55", "").replace("55", "")}' if pd.notna(x) and x != '<NA>' else '')

            if 'Numero' in df_export.columns:
                df_export['Numero'] = pd.to_numeric(df_export['Numero'], errors='coerce')
                df_export['Numero'] = df_export['Numero'].astype('Int64').astype(str)
                df_export['Numero'] = df_export['Numero'].apply(lambda x: str(x) if pd.notna(x) and x != '<NA>' else '')

            st.session_state.df_export = df_export

            st.subheader("Opções de Exportação")
            if "filename" not in st.session_state:
                st.session_state.filename = f"relatorio_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

            filename_input = st.text_input("Nome do arquivo (sem extensão)", value=st.session_state.filename, key="filename_input_key")
            st.session_state.filename = filename_input

            pdf_title_input = st.text_input("Título do PDF", value="Relatório de Clientes", key="pdf_title_input_key")
            st.session_state.pdf_title = pdf_title_input

            current_date = datetime.now().strftime('%d-%m-%Y')
            final_output_filename = f"{st.session_state.filename}_{current_date}"

            col1, col2 = st.columns(2)
            with col1:
                if st.button("Gerar e Baixar PDF"):
                    with st.spinner("Gerando PDF..."):
                        pdf_buffer = create_pdf_robust(st.session_state.df_export, title=st.session_state.pdf_title)
                        if pdf_buffer:
                            st.session_state.pdf_buffer = pdf_buffer
                            st.session_state.pdf_filename = final_output_filename + ".pdf"
                        else:
                            st.error("Falha ao gerar o PDF.")
            
            with col2:
                if st.button("Gerar e Baixar Excel (XLSX)"):
                    with st.spinner("Gerando Excel..."):
                        output = io.BytesIO()
                        st.session_state.df_export.to_excel(output, index=False)
                        output.seek(0)
                        st.session_state.excel_buffer = output
                        st.session_state.excel_filename = final_output_filename + ".xlsx"

            if 'pdf_buffer' in st.session_state and st.session_state.pdf_buffer:
                st.download_button(
                    label="Baixar PDF Gerado",
                    data=st.session_state.pdf_buffer,
                    file_name=st.session_state.pdf_filename,
                    mime="application/pdf",
                    key='download_pdf_higienizacao'
                )
                # Limpa o buffer após o botão ser exibido
                # st.session_state.pdf_buffer = None 

            if 'excel_buffer' in st.session_state and st.session_state.excel_buffer:
                st.download_button(
                    label="Baixar XLSX Gerado",
                    data=st.session_state.excel_buffer,
                    file_name=st.session_state.excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key='download_excel_higienizacao'
                )
                # Limpa o buffer após o botão ser exibido
                # st.session_state.excel_buffer = None

def aba_divisor_listas():
    st.header("Divisor de Listas de Leads")
    uploaded_file = st.file_uploader("Faça upload do arquivo XLSX com os leads", type=["xlsx"], key="divisor_uploader")
    
    if uploaded_file:
        # Load raw data immediately after upload to get columns for mapping
        df_raw_leads, err = load_data(uploaded_file)
        if err:
            st.error(err)
            return

        st.subheader("Opções de Filtragem e Distribuição")
        
        col1, col2 = st.columns(2)

        with col1:
            # Filtro por data de início
            start_date = st.date_input("Data de Início da Distribuição", value=date.today(), help="Selecione a data a partir da qual a distribuição de leads começará.")

        with col2:
            # Filtro por equipe/supervisor
            all_teams = list(EQUIPES.keys())
            selected_teams = st.multiselect(
                "Filtrar por Equipe/Supervisor", 
                options=all_teams, 
                default=all_teams,
                help="Selecione as equipes cujos consultores devem receber leads. Se nenhuma for selecionada, todos os consultores serão considerados inicialmente.",
                key="divisor_filter_teams"
            )

        # Filtrar consultores a serem excluídos (mantido abaixo para melhor visualização de muitas opções)
        consultants_pool = []
        if selected_teams:
            for team in selected_teams:
                consultants_pool.extend(EQUIPES.get(team, []))
            consultants_pool = sorted(list(set(consultants_pool)))
        else:
            consultants_pool = sorted(CONSULTORES)

        excluded_consultants = st.multiselect(
            "Excluir Consultores Específicos", 
            options=consultants_pool,
            help="Selecione os consultores que NÃO devem receber leads nesta distribuição.",
            key="divisor_exclude_consultants"
        )

        leads_per_consultant = st.number_input("Quantidade de leads por consultor", min_value=1, value=50, help="Defina quantos leads cada consultor receberá por vez.")

        st.subheader("Mapeamento de Colunas de Entrada")
        st.info("O sistema tentará mapear as colunas 'NOME' e 'Whats' automaticamente. Verifique e ajuste se necessário.")

        df_leads_cols = df_raw_leads.columns.tolist()
        expected_cols_divisor = ["NOME", "Whats", "CEL"]
        
        # Sugestões de nomes de colunas para pré-seleção automática
        SUGGESTED_COLUMN_NAMES = {
            "NOME": ["NOME", "Nome Completo", "Cliente", "Razao Social", "Empresa", "NOME/RAZAO_SOCIAL", "Socio1Nome", "Nome", "Razao"],
            "Whats": ["Whats", "WhatsApp", "Telefone", "Celular", "Contato", "CELULAR1", "SOCIO1Celular1", "Socio1Celular1"],
            "CEL": ["CEL", "Celular", "Telefone", "Whats", "WhatsApp", "CELULAR2", "SOCIO1Celular2", "Socio1Celular2"]
        }

        user_col_mapping = {}
        # Mapeia as colunas do arquivo para minúsculas para busca case-insensitive
        df_cols_lower_map = {c.lower(): c for c in reversed(df_leads_cols)}

        for col in expected_cols_divisor:
            default_selection = ''
            
            # A lista de busca prioriza o nome exato da coluna esperada, depois as sugestões
            search_list = [col] + SUGGESTED_COLUMN_NAMES.get(col, [])

            for suggested_name in search_list:
                # Busca case-insensitive pela sugestão no mapeamento de colunas do arquivo
                if suggested_name.lower() in df_cols_lower_map:
                    default_selection = df_cols_lower_map[suggested_name.lower()]
                    break
            
            # Determina o índice da opção pré-selecionada para o selectbox
            try:
                # Adiciona 1 porque a lista de opções do selectbox começa com um item vazio ''
                default_index = df_leads_cols.index(default_selection) + 1 if default_selection else 0
            except ValueError:
                default_index = 0

            selected_col = st.selectbox(
                f"Coluna para '{col}'",
                options=[''] + df_leads_cols,
                index=default_index,
                key=f"map_divisor_{col}"
            )
            user_col_mapping[col] = selected_col
        
        if st.button("Processar e Gerar Listas"):
            with st.spinner("Processando... Por favor, aguarde."):
                try:
                    # Validate NOME mapping before proceeding
                    if not user_col_mapping["NOME"]:
                        st.warning("A coluna 'NOME' é obrigatória para a distribuição de leads.")
                        return

                    # Apply mapping and rename DataFrame
                    df_leads_mapped = df_raw_leads.copy()
                    for expected, actual in user_col_mapping.items():
                        if actual: # Only process if a column was selected
                            if actual in df_leads_mapped.columns:
                                df_leads_mapped.rename(columns={actual: expected}, inplace=True)
                            else:
                                st.warning(f"A coluna '{actual}' selecionada para '{expected}' não foi encontrada no arquivo. Verifique o mapeamento.")
                                return

                    # Validate if NOME column exists after mapping
                    if "NOME" not in df_leads_mapped.columns:
                        st.warning("A coluna 'NOME' é obrigatória para a distribuição de leads e não foi mapeada corretamente.")
                        return

                    # Limpa e filtra pelo número de WhatsApp
                    if "Whats" in df_leads_mapped.columns:
                        initial_rows = len(df_leads_mapped)
                        df_leads_mapped["Whats"] = df_leads_mapped["Whats"].apply(clean_phone_number)
                        df_leads_mapped.dropna(subset=["Whats"], inplace=True)
                        final_rows = len(df_leads_mapped)
                        st.info(f"{initial_rows - final_rows} linhas foram removidas por não conterem um número de WhatsApp válido.")
                    else:
                        st.warning("A coluna 'Whats' não foi mapeada. Nenhuma filtragem por WhatsApp foi aplicada.")

                    if df_leads_mapped.empty:
                        st.warning("Após a filtragem, não restaram leads para distribuir.")
                        return

                    # Determine effective consultants based on filters
                    effective_consultores = []
                    if selected_teams:
                        for team in selected_teams:
                            effective_consultores.extend(EQUIPES.get(team, []))
                        effective_consultores = list(set(effective_consultores))
                    else:
                        effective_consultores = list(CONSULTORES)

                    if excluded_consultants:
                        effective_consultores = [c for c in effective_consultores if c not in excluded_consultants]
                    effective_consultores.sort()

                    if not effective_consultores:
                        st.warning("Nenhum consultor selecionado após a aplicação dos filtros. Ajuste suas seleções.")
                        return

                    # Clean CEL column (now in df_leads_mapped)
                    if "CEL" in df_leads_mapped.columns:
                        df_leads_mapped["CEL"] = pd.to_numeric(df_leads_mapped["CEL"], errors='coerce')
                        df_leads_mapped["CEL"] = df_leads_mapped["CEL"].astype('Int64').astype(str).replace('<NA>', '')
                    
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                        leads_processados = 0
                        data_atual = start_date
                        total_leads = len(df_leads_mapped)
                        arquivos_gerados = 0

                        while leads_processados < total_leads:
                            for consultor in effective_consultores:
                                if leads_processados >= total_leads: 
                                    break

                                inicio_lote = leads_processados
                                fim_lote = leads_processados + leads_per_consultant
                                st.info(f"Processando leads de {inicio_lote} a {fim_lote} para o consultor {consultor}")
                                df_lote = df_leads_mapped.iloc[inicio_lote:fim_lote].copy()

                                # Convert numeric columns to string
                                for col in df_lote.columns:
                                    if pd.api.types.is_numeric_dtype(df_lote[col]):
                                        df_lote[col] = df_lote[col].astype('Int64').astype(str).replace('<NA>', '')

                                # Define and ensure checkbox columns
                                cols_to_center = ["1º Contato", "2º Contato", "3º Contato", "Atend. Lig.(S/N)", "Visita Marc.(S/N)"]
                                cols_single_checkbox = ["1º Contato", "2º Contato", "3º Contato"]
                                cols_double_checkbox = ["Atend. Lig.(S/N)", "Visita Marc.(S/N)"]

                                for col in cols_single_checkbox:
                                    if col not in df_lote.columns:
                                        df_lote[col] = "☐"
                                    else:
                                        df_lote[col] = "☐"
                                
                                for col in cols_double_checkbox:
                                    if col not in df_lote.columns:
                                        df_lote[col] = "☐   ☐"
                                    else:
                                        df_lote[col] = "☐   ☐"
                                
                                if not df_lote.empty:
                                    excel_buffer = gerar_excel_em_memoria(df_lote, consultor, data_atual)
                                    
                                    primeiro_nome = consultor.split(' ')[0]
                                    data_formatada_nome = data_atual.strftime('%d_%m_%Y')
                                    nome_arquivo_base = f"LEADS_AUTOMOVEIS_{primeiro_nome.upper()}_{data_formatada_nome}"
                                    
                                    nome_equipe = CONSULTOR_PARA_EQUIPE.get(consultor, "Outros")
                                    zip_file.writestr(f"{nome_equipe}/{nome_arquivo_base}.xlsx", excel_buffer.getvalue())

                                    pdf_title = f"Leads Automoveis - {primeiro_nome} {data_atual.strftime('%d/%m')}"
                                    pdf_buffer = create_pdf_robust(df_lote, title=pdf_title, cols_to_center=cols_to_center, cols_single_checkbox=cols_single_checkbox, cols_double_checkbox=cols_double_checkbox)
                                    
                                    if pdf_buffer:
                                        zip_file.writestr(f"{nome_equipe}/{nome_arquivo_base}.pdf", pdf_buffer.getvalue())
                                    
                                    leads_processados += len(df_lote)
                                    arquivos_gerados += 1

                            data_atual = proximo_dia_util(data_atual)
                    
                    st.success(f"Processo concluído! {arquivos_gerados} pares de listas (Excel e PDF) foram gerados.")
                    
                    zip_filename = f"Listas_Consultores_{datetime.now().strftime('%d-%m-%Y')}.zip"
                    st.download_button(
                        label="Baixar Todas as Listas (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name=zip_filename,
                        mime="application/zip"
                    )

                except Exception as e:
                    st.error(f"Ocorreu um erro durante o processamento: {e}")

def aba_gerador_negocios_robos():
    st.header("Gerador de Negócios para Robôs")

    st.info("Esta aba gera planilhas de 'Negócios' para importação em sistemas de robôs, utilizando os arquivos de 'Pessoas Agendor' gerados anteriormente.")

    input_folder_path = st.text_input("Caminho da pasta com os arquivos de 'Pessoas Agendor' (ex: pessoas_geradas)", value="pessoas_geradas", help="Insira o caminho da pasta que contém os arquivos XLSX de 'Pessoas Agendor'. O script buscará em subpastas também.")

    if input_folder_path:
        st.subheader("Configurações para Geração de Negócios")

        col1, col2 = st.columns(2)
        with col1:
            negocios_por_consultor = st.number_input("Número de negócios por consultor (por arquivo)", min_value=1, value=20, help="Define quantos leads serão incluídos em cada arquivo de negócio gerado para um consultor.")
        with col2:
            start_date_negocios = st.date_input("Data de Início para Negócios", value=date.today(), help="A data de início para o primeiro arquivo de negócio. As datas subsequentes serão incrementadas, pulando finais de semana.")

        col3, col4 = st.columns(2)
        with col3:
            nicho_principal = st.text_input("Nicho Principal (ex: AUTO, MED, EMPR)", value="AUTO", help="O nicho principal para o 'Título do negócio'.")
        with col4:
            sufixo_localidade = st.text_input("Sufixo de Localidade (opcional, ex: CG, MS)", value="", help="Um sufixo opcional para o nicho, como a localidade.")
        
        if st.button("Gerar Arquivos de Negócios"):
            if not input_folder_path:
                st.warning("Por favor, insira o caminho da pasta com os arquivos de 'Pessoas Agendor'.")
                return

            if not nicho_principal:
                st.error("Por favor, insira o Nicho Principal.")
                return

            with st.spinner("Gerando arquivos de Negócios... Por favor, aguarde."):
                all_generated_files = {}

                # Construir o caminho absoluto para a pasta de entrada
                project_root = os.getcwd()
                absolute_input_path = os.path.join(project_root, input_folder_path)

                # Verificar se a pasta existe
                if not os.path.isdir(absolute_input_path):
                    st.error(f"A pasta especificada não foi encontrada: {absolute_input_path}")
                    return

                # Encontrar todos os arquivos XLSX na pasta e subpastas
                all_pessoas_files = glob.glob(os.path.join(absolute_input_path, "**/*.xlsx"), recursive=True)

                if not all_pessoas_files:
                    st.warning(f"Nenhum arquivo XLSX encontrado na pasta: {absolute_input_path}")
                    return

                for file_path in all_pessoas_files:
                    try:
                        df_pessoas, err = load_data(file_path)
                        if err:
                            st.warning(f"Erro ao carregar {os.path.basename(file_path)}: {err}. Pulando este arquivo.")
                            continue
                        
                        # Extrair o nome do consultor do nome do arquivo de pessoas
                        # Ex: PESSOAS_GERAL_CG_CONSULTOR_NOME_16-07-2025.xlsx
                        file_name_only = os.path.basename(file_path)
                        file_name_parts = file_name_only.replace(".xlsx", "").split('_')
                        consultor_nome_arquivo = ""
                        if len(file_name_parts) >= 4:
                            consultor_nome_arquivo = file_name_parts[3] # Pega o nome do consultor
                        
                        if not consultor_nome_arquivo:
                            st.warning(f"Não foi possível extrair o nome do consultor do arquivo: {file_name_only}. Pulando este arquivo.")
                            continue

                        # Colunas da planilha de Negócios
                        colunas_negocios = [
                            "Título do negócio", "Empresa relacionada", "Pessoa relacionada",
                            "Usuário responsável", "Data de início", "Data de conclusão",
                            "Valor Total", "Funil", "Etapa", "Status", "Motivo de perda",
                            "Descrição do motivo de perda", "Ranking", "Descrição", "Produtos e Serviços"
                        ]

                        leads_do_consultor = df_pessoas.copy()
                        
                        # Garantir que as colunas essenciais existam
                        required_cols_pessoas = ["Nome", "Usuário responsável", "WhatsApp"]
                        if not all(col in leads_do_consultor.columns for col in required_cols_pessoas):
                            st.warning(f"Arquivo {file_name_only} não contém todas as colunas essenciais (Nome, Usuário responsável, WhatsApp). Pulando este arquivo.")
                            continue

                        # Limpar e formatar WhatsApp para uso em Data de Conclusão
                        leads_do_consultor["WhatsApp_Clean"] = leads_do_consultor["WhatsApp"].apply(clean_phone_number)
                        leads_do_consultor["WhatsApp_Clean"] = leads_do_consultor["WhatsApp_Clean"].apply(lambda x: str(int(x)) if pd.notna(x) else "")

                        num_leads_consultor = len(leads_do_consultor)
                        leads_processados_consultor = 0
                        current_date = start_date_negocios
                        file_counter = 1

                        while leads_processados_consultor < num_leads_consultor:
                            inicio_lote = leads_processados_consultor
                            fim_lote = min(leads_processados_consultor + negocios_por_consultor, num_leads_consultor)
                            df_lote_negocios = leads_do_consultor.iloc[inicio_lote:fim_lote].copy()

                            if not df_lote_negocios.empty:
                                dados_negocios = []
                                for _, row_lead in df_lote_negocios.iterrows():
                                    nome_pessoa = row_lead.get("Nome", "")
                                    usuario_responsavel = row_lead.get("Usuário responsável", "")
                                    whatsapp_lead = row_lead.get("WhatsApp_Clean", "")

                                    # Formatar Título do negócio
                                    mes_ano = datetime.now().strftime('%m/%y')
                                    nicho_formatado_titulo = nicho_principal.upper()
                                    if sufixo_localidade:
                                        nicho_formatado_titulo += f" {sufixo_localidade.upper()}"
                                    
                                    titulo_negocio = f"{mes_ano} - RB - {nicho_formatado_titulo} - {nome_pessoa}/ESPs"

                                    linha_negocio = {
                                        "Título do negócio": titulo_negocio,
                                        "Empresa relacionada": "", # Deixar em branco
                                        "Pessoa relacionada": nome_pessoa,
                                        "Usuário responsável": usuario_responsavel,
                                        "Data de início": current_date.strftime('%d/%m/%Y'),
                                        "Data de conclusão": whatsapp_lead, # WhatsApp temporariamente aqui
                                        "Valor Total": "", # Deixar em branco
                                        "Funil": "Funil de Vendas",
                                        "Etapa": "Prospecção",
                                        "Status": "Em andamento",
                                        "Motivo de perda": "", # Deixar em branco
                                        "Descrição do motivo de perda": "", # Deixar em branco
                                        "Ranking": "", # Deixar em branco
                                        "Descrição": "", # Deixar em branco
                                        "Produtos e Serviços": "" # Deixar em branco
                                    }
                                    dados_negocios.append(linha_negocio)
                                
                                df_final_negocios = pd.DataFrame(dados_negocios, columns=colunas_negocios)

                                output_excel_negocios = io.BytesIO()
                                with pd.ExcelWriter(output_excel_negocios, engine='openpyxl') as writer:
                                    df_final_negocios.to_excel(writer, index=False)
                                output_excel_negocios.seek(0)

                                # Nome do arquivo de negócios
                                nome_arquivo_negocios = f"NEGOCIOS_{consultor_nome_arquivo.upper()}_{nicho_principal.upper()}"
                                if sufixo_localidade:
                                    nome_arquivo_negocios += f"_{sufixo_localidade.upper()}"
                                nome_arquivo_negocios += f"_{current_date.strftime('%d-%m-%Y')}.xlsx"
                                
                                # Obter o caminho relativo da pasta do consultor para o ZIP
                                # Ex: pessoas_geradas/Equipe Camila/PESSOAS_... -> Equipe Camila/
                                relative_path_parts = os.path.relpath(os.path.dirname(file_path), absolute_input_path).split(os.sep)
                                consultor_zip_folder = "" # Default para o root do zip
                                if len(relative_path_parts) > 0 and relative_path_parts[0] != '.': # Se não for o root da pasta de input
                                    consultor_zip_folder = os.path.join(*relative_path_parts) + os.sep

                                all_generated_files[os.path.join(consultor_zip_folder, nome_arquivo_negocios)] = output_excel_negocios.getvalue()

                                leads_processados_consultor += len(df_lote_negocios)
                                current_date = proximo_dia_util(current_date) # Avança a data para o próximo arquivo
                                file_counter += 1

                    except Exception as e:
                        st.error(f"Erro ao processar o arquivo {os.path.basename(file_path)}: {e}")
                        continue
                
                if all_generated_files:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                        for file_name_in_zip, file_data in all_generated_files.items():
                            zip_file.writestr(file_name_in_zip, file_data)
                    
                    zip_filename = f"Negocios_Robos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
                    st.download_button(
                        label="Baixar Todos os Arquivos de Negócios (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name=zip_filename,
                        mime="application/zip"
                    )
                    st.success(f"Processo concluído! {len(all_generated_files)} arquivos de Negócios gerados.")
                else:
                    st.warning("Nenhum arquivo de Negócios foi gerado. Verifique os arquivos de entrada e as configurações.")


def aba_automacao_pessoas_agendor():
    st.header("Automação Pessoas Agendor")

    uploaded_file = st.file_uploader("Faça upload do arquivo XLSX com os leads", type=["xlsx"], key="geracao_pessoas_uploader")

    if uploaded_file:
        df_raw_leads, err = load_data(uploaded_file)
        if err:
            st.error(err)
            return

        st.subheader("Opções de Filtragem e Distribuição")
        
        # Filtro por equipe/supervisor
        all_teams = list(EQUIPES.keys())
        selected_teams = st.multiselect(
            "Filtrar por Equipe/Supervisor", 
            options=all_teams, 
            default=all_teams,
            help="Selecione as equipes cujos consultores devem receber leads. Se nenhuma for selecionada, todos os consultores serão considerados inicialmente.",
            key="agendor_filter_teams"
        )

        # Filtrar consultores a serem excluídos
        consultants_pool = []
        if selected_teams:
            for team in selected_teams: 
                consultants_pool.extend(EQUIPES.get(team, []))
            consultants_pool = sorted(list(set(consultants_pool)))
        else:
            consultants_pool = sorted(CONSULTORES)

        excluded_consultants = st.multiselect(
            "Excluir Consultores Específicos", 
            options=consultants_pool,
            help="Selecione os consultores que NÃO devem receber leads nesta distribuição.",
            key="agendor_exclude_consultants"
        )

        st.subheader("Configurações Adicionais para Agendor")
        default_cargo = st.text_input("Cargo Padrão", value="Lead Automovel", help="Cargo a ser atribuído aos leads no Agendor.")
        default_descricao = st.text_area("Descrição Padrão", value="", help="Descrição adicional para os leads no Agendor.")
        default_uf = st.text_input("UF Padrão", value="MS", max_chars=2, help="UF padrão para os leads, se não mapeado.")
        nicho_valor = st.text_input("Nicho (para nome do arquivo)", value="GERAL", help="Valor do nicho para o nome do arquivo de exportação (ex: AUTOMOVEIS, IMOVEIS).")

        st.subheader("Mapeamento de Colunas")
        st.info("Selecione as colunas do seu arquivo que correspondem aos campos esperados.")

        df_leads_cols = df_raw_leads.columns.tolist()
        
        # Suggested column names for pre-selection
        SUGGESTED_COLUMN_NAMES_AGENDOR = {
            "NOME": ["NOME", "Nome Completo", "Socio1Nome", "Razao Social", "Empresa", "NOME/RAZAO_SOCIAL"],
            "Whats": ["Whats", "WhatsApp", "Telefone", "Celular", "Contato", "CELULAR1", "CELULAR2", "SOCIO1Celular1"],
            "CEL": ["CEL", "Celular", "Telefone", "Whats", "WhatsApp", "CELULAR1", "CELULAR2", "SOCIO1Celular2"],
            "Rua": ["Rua", "Logradouro", "Endereco", "Endereço"],
            "Número": ["Numero", "Número", "Num"],
            "Bairro": ["Bairro"],
            "Cidade": ["Cidade"]
        }

        expected_cols_agendor = ["NOME", "Whats", "CEL", "Rua", "Número", "Bairro", "Cidade"]
        user_col_mapping = {}
        # Mapeia as colunas do arquivo para minúsculas para busca case-insensitive
        df_cols_lower_map = {c.lower(): c for c in reversed(df_leads_cols)}

        for col in expected_cols_agendor:
            default_selection = ''
            
            # A lista de busca prioriza o nome exato da coluna esperada, depois as sugestões
            search_list = [col] + SUGGESTED_COLUMN_NAMES_AGENDOR.get(col, [])

            for suggested_name in search_list:
                # Busca case-insensitive pela sugestão no mapeamento de colunas do arquivo
                if suggested_name.lower() in df_cols_lower_map:
                    default_selection = df_cols_lower_map[suggested_name.lower()]
                    break
            
            # Determina o índice da opção pré-selecionada para o selectbox
            try:
                # Adiciona 1 porque a lista de opções do selectbox começa com um item vazio ''
                default_index = df_leads_cols.index(default_selection) + 1 if default_selection else 0
            except ValueError:
                default_index = 0

            selected_col = st.selectbox(
                f"Coluna para '{col}'",
                options=[''] + df_leads_cols,
                index=default_index,
                key=f"map_agendor_{col}"
            )
            user_col_mapping[col] = selected_col

        leads_por_consultor = st.number_input("Número de leads por consultor", min_value=1, value=50)

        if st.button("Gerar Arquivo 'Pessoas'"):
            with st.spinner("Processando... Por favor, aguarde."):
                try:
                    # Validate NOME mapping before proceeding
                    if not user_col_mapping["NOME"]:
                        st.warning("A coluna 'NOME' é obrigatória para a distribuição de leads.")
                        return

                    # Apply mapping and rename DataFrame
                    df_leads_mapped = df_raw_leads.copy()
                    for expected, actual in user_col_mapping.items():
                        if actual: # Only process if a column was selected
                            if actual in df_leads_mapped.columns:
                                df_leads_mapped.rename(columns={actual: expected}, inplace=True)
                            else:
                                st.warning(f"A coluna '{actual}' selecionada para '{expected}' não foi encontrada no arquivo. Verifique o mapeamento.")
                                return

                    # Validate if NOME column exists after mapping
                    if "NOME" not in df_leads_mapped.columns:
                        st.warning("A coluna 'NOME' é obrigatória para a distribuição de leads e não foi mapeada corretamente.")
                        return

                    # Limpa e filtra pelo número de WhatsApp
                    if "Whats" in df_leads_mapped.columns:
                        initial_rows = len(df_leads_mapped)
                        df_leads_mapped["Whats"] = df_leads_mapped["Whats"].apply(clean_phone_number)
                        df_leads_mapped.dropna(subset=["Whats"], inplace=True)
                        final_rows = len(df_leads_mapped)
                        st.info(f"{initial_rows - final_rows} linhas foram removidas por não conterem um número de WhatsApp válido.")
                    else:
                        st.warning("A coluna 'Whats' não foi mapeada. Nenhuma filtragem por WhatsApp foi aplicada.")

                    if df_leads_mapped.empty:
                        st.warning("Após a filtragem, não restaram leads para distribuir.")
                        return

                    # Determine effective consultants based on filters
                    effective_consultores = []
                    if selected_teams:
                        for team in selected_teams:
                            effective_consultores.extend(EQUIPES.get(team, []))
                        effective_consultores = list(set(effective_consultores))
                    else:
                        effective_consultores = list(CONSULTORES)

                    if excluded_consultants:
                        effective_consultores = [c for c in effective_consultores if c not in excluded_consultants]
                    effective_consultores.sort()

                    if not effective_consultores:
                        st.warning("Nenhum consultor selecionado após a aplicação dos filtros. Ajuste suas seleções.")
                        return

                    # Clean CEL column (now in df_leads_mapped)
                    if "CEL" in df_leads_mapped.columns:
                        df_leads_mapped["CEL"] = pd.to_numeric(df_leads_mapped["CEL"], errors='coerce')
                        df_leads_mapped["CEL"] = df_leads_mapped["CEL"].astype('Int64').astype(str).replace('<NA>', '')
                    
                    # --- Agendor Specific Logic ---
                    # Deduplicate by WhatsApp
                    if "Whats" in df_leads_mapped.columns:
                        df_leads_mapped.drop_duplicates(subset=["Whats"], keep='first', inplace=True)
                        st.info(f"Leads após desduplicação por WhatsApp: {len(df_leads_mapped)}")

                    # Prepare for Agendor output
                    colunas_output = [
                        "Nome", "CPF", "Empresa", "Cargo", "Aniversário", "Ano de nascimento", 
                        "Usuário responsável", "Categoria", "Origem", "Descrição", "E-mail", 
                        "WhatsApp", "Telefone", "Celular", "Fax", "Ramal", "CEP", "País", 
                        "Estado", "Cidade", "Bairro", "Rua", "Número", "Complemento", 
                        "Produto", "Facebook", "Twitter", "LinkedIn", "Skype", "Instagram", "Ranking"
                    ]
                    
                    # --- Lógica de Geração e Download ---
                    # Armazena os arquivos gerados em memória
                    generated_files = {}

                    leads_processados = 0
                    total_leads = len(df_leads_mapped)

                    st.info(f"DEBUG: df_leads_mapped length after all filters: {len(df_leads_mapped)}")
                    st.info(f"DEBUG: Valor de leads_por_consultor ANTES do loop de distribuição: {leads_por_consultor}")

                    while leads_processados < total_leads:
                        for consultor in effective_consultores:
                            if leads_processados >= total_leads:
                                break

                            inicio_lote = leads_processados
                            fim_lote = leads_processados + leads_por_consultor
                            df_lote = df_leads_mapped.iloc[inicio_lote:fim_lote].copy()

                            if not df_lote.empty:
                                # ... (código de preparação do df_final_consultor) ...
                                dados_finais = []
                                consultor_formatado = consultor.lower().replace(' ', '.')
                                for _, row in df_lote.iterrows():
                                        # Limpa o número e adiciona o prefixo apenas se houver um valor válido
                                        whatsapp_val = row.get("Whats")
                                        whatsapp_str = f"+55{str(whatsapp_val).split('.')[0]}" if whatsapp_val and pd.notna(whatsapp_val) else ""

                                        celular_val = row.get("CEL")
                                        celular_str = f"+55{str(celular_val).split('.')[0]}" if celular_val and pd.notna(celular_val) else ""

                                        linha = {col: "" for col in colunas_output}
                                        linha.update({
                                        "Nome": row.get("NOME", ""),
                                        "Cargo": default_cargo,
                                        "Usuário responsável": consultor_formatado,
                                        "Categoria": "Lead",
                                        "Origem": "Reobote",
                                        "Descrição": default_descricao,
                                        "WhatsApp": whatsapp_str,
                                        "Celular": celular_str,
                                        "Estado": default_uf,
                                        "Cidade": row.get("Cidade", ""),
                                        "Bairro": row.get("Bairro", ""),
                                        "Rua": row.get("Rua", ""),
                                        "Número": row.get("Número", ""),
                                        "Complemento": row.get("Complemento", "")
                                    })
                                        dados_finais.append(linha)
                                
                                df_final_consultor = pd.DataFrame(dados_finais, columns=colunas_output)

                                output_excel_consultor = io.BytesIO()
                                with pd.ExcelWriter(output_excel_consultor, engine='openpyxl') as writer:
                                    df_final_consultor.to_excel(writer, index=False, sheet_name='Pessoas')
                                output_excel_consultor.seek(0)

                                # Lógica para determinar a localidade
                                localidade = "CG" # Padrão
                                cidade_col = user_col_mapping.get("Cidade")
                                uf_col = user_col_mapping.get("UF")
                                if cidade_col and cidade_col in df_lote.columns and not df_lote[cidade_col].empty:
                                    cidade_val = df_lote[cidade_col].iloc[0]
                                    if pd.notna(cidade_val) and str(cidade_val).strip():
                                        localidade = str(cidade_val).strip().upper()
                                elif uf_col and uf_col in df_lote.columns and not df_lote[uf_col].empty:
                                    uf_val = df_lote[uf_col].iloc[0]
                                    if pd.notna(uf_val) and str(uf_val).strip():
                                        localidade = str(uf_val).strip().upper()

                                # Formata o nicho e o nome do consultor para o nome do arquivo
                                nicho_formatado = nicho_valor.upper().replace(' ', '_')
                                primeiro_nome = consultor.split(' ')[0].upper()
                                data_formatada = datetime.now().strftime('%d-%m-%Y')
                                
                                nome_arquivo_agendor = f"PESSOAS_{nicho_formatado}_{localidade}_{primeiro_nome}_{data_formatada}.xlsx"
                                
                                generated_files[nome_arquivo_agendor] = output_excel_consultor.getvalue()
                                
                                leads_processados += len(df_lote)

                    # --- Lógica de Download (Inteligente) ---
                    if not generated_files:
                        st.warning("Nenhum arquivo foi gerado. Verifique os filtros e os dados de entrada.")
                        return

                    st.success(f"Processo concluído! {len(generated_files)} arquivo(s) de pessoas para Agendor foram gerados.")

                    # Se for apenas um arquivo, oferece o download direto
                    if len(generated_files) == 1:
                        file_name, file_data = list(generated_files.items())[0]
                        st.download_button(
                            label=f"Baixar Arquivo para Agendor (.xlsx)",
                            data=file_data,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    # Se forem vários arquivos, agrupa em um ZIP
                    else:
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                            for file_name, file_data in generated_files.items():
                                # Extrai o nome do consultor do nome do arquivo para encontrar a equipe
                                parts = file_name.split('_')
                                consultor_nome_no_arquivo = ""
                                if len(parts) > 3:
                                    consultor_nome_no_arquivo = parts[-2].upper()

                                nome_equipe = "Outros" # Padrão
                                for consultor_lista, equipe in CONSULTOR_PARA_EQUIPE.items():
                                    if consultor_lista.split(' ')[0].upper() == consultor_nome_no_arquivo:
                                        nome_equipe = equipe
                                        break
                                zip_file.writestr(f"{nome_equipe}/{file_name}", file_data)
                        
                        zip_filename = f"Pessoas_Agendor_Distribuicao_{datetime.now().strftime('%d-%m-%Y')}.zip"
                        st.download_button(
                            label="Baixar Todos os Arquivos para Agendor (ZIP)",
                            data=zip_buffer.getvalue(),
                            file_name=zip_filename,
                            mime="application/zip"
                        )

                except Exception as e:
                    st.error(f"Ocorreu um erro durante o processamento: {e}")


def main():
    st.title("Automação de Listas")

    tabs = st.tabs(["Higienização de Dados", "Divisor de Listas Diárias", "Robos Lista Automoveis", "Automação Pessoas Agendor"])

    with tabs[0]:
        aba_higienizacao()

    with tabs[1]:
        aba_divisor_listas()
    
    with tabs[2]:
        aba_gerador_negocios_robos()

    with tabs[3]:
        aba_automacao_pessoas_agendor()

if __name__ == "__main__":
    main()
